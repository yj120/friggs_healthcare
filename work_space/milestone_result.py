from openpyxl import load_workbook
import os
import scipy.stats as stats
import numpy as np

load_file_name= 'C:/Users/홍예지/Desktop/프릭스헬스케어/survival_function.xlsx'

load_wb=load_workbook(os.path.join(os.getcwd(),load_file_name))
load_ws=load_wb['Sheet1']

milestone = int(input("====> 마일스톤 넘버 입력 : "))
day = int(input("====> 일자 입력 : "))


def milestone_info(milestone, day):
    print(milestone, "번의 평균 가능 일자 :", load_ws.cell(milestone + 1, 2).value,"일")
    average = load_ws.cell(milestone + 1, 2).value
    std = load_ws.cell(milestone + 1, 3).value
    start_day = load_ws.cell(milestone + 1, 4).value
    due_day = load_ws.cell(milestone + 1, 5).value
    rv = stats.norm(average, std)
    x = np.linspace(start_day, due_day, average)
    y2 = rv.pdf(x)
    y3 = rv.cdf(day)

    print(milestone, "번은 상위", "%.2f%%" % (y3 * 100), "입니다.")
    return


milestone_info(milestone, day)