
# 1. Data_preprocessing class : 전처리
# 2. AverageDate class : kaplan-meier 를 이용한 평균일자 계산
# 3. Update class : 업데이트(데이터 셋 병합) * 업데이트 시 사용
# 4. final_result function : 최종 결과 확인

#================================================================================================================

import pandas as pd
import numpy as np

df_refine_astype = pd.DataFrame()  # 0시 제거 시 사용된 변수
df_temporary_1 = pd.DataFrame() # created_at 에서 시간 저장 변수
df_temporary_1 = pd.DataFrame() # updated_at 에서 시간 저장 변수
result = pd.DataFrame() # created_at 시간이 00:00:00.000 을 뺀 데이터 셋
milestone_content_uniquenumber = pd.DataFrame() # unique_number 매핑 파일 담을 변수
df_unique_mapping = pd.DataFrame() # unique_number mapping 시킨 데이터셋
gender_set = pd.DataFrame() # gender 매핑 파일 담을 변수
gender_info_mapping = pd.DataFrame() # gender mapping 시킨 데이터셋
condition_1 = pd.DataFrame() # done_day_after_birth 기간 조건(시작일)
condition_2 = pd.DataFrame() # done_day_after_birth 기간 조건(종료일)
df_date_range = pd.DataFrame() # condition_1 과 condition_2 기간 조건이 적용된 데이터셋
condition_3 = pd.DataFrame() # gender == male 인 데이터셋
condition_4 = pd.DataFrame() # gender == female 인 데이터셋
condition_5 = pd.DataFrame() # gender == both(male+female) 데이터셋
range_input = 0 # milestone_unique 최대값


class Data_preprocessing: # 데이터 전처리 클래스

    def __init__(self, df='None'):
        self.df = pd.read_excel(df)

    def print_dataset(self): # 데이터셋 확인
        return self.df

    def duplicates(self): # 중복 제거
        self.df = self.df.drop_duplicates()
        return self.df

    def delete_zero_time(self): # 0시 제거
        df_refine_astype = self.df.astype({'created_at': 'str', 'updated_at': 'str'})
        df_temporary_1 = pd.DataFrame()
        df_temporary_2 = pd.DataFrame()
        # df_temporary_1 = df_temporary_1.copy()
        # df_temporary_2 = df_temporary_2.copy()
        df_temporary_1 = df_refine_astype['created_at'].str.split(" ")
        df_temporary_2 = df_refine_astype['updated_at'].str.split(" ")
        df_temporary_1 = df_temporary_1.str[-1]
        df_temporary_2 = df_temporary_2.str[-1]
        result = pd.concat([self.df, df_temporary_1, df_temporary_2], axis=1)
        result.columns = ['person', 'milestone', 'grade', 'done_day_after_birth', 'created_at', 'updated_at',
                          'created_hour', 'updated_hour']
        self.df = result[result.created_hour != '00:00:00.000']
        return self.df

    def unique_number_mapping(self):  # unique_number_mapping only
        milestone_content_uniquenumber = pd.read_excel(
            'C:/Users/홍예지/Desktop/프릭스헬스케어/milestone_content_uniquenumber_test_ver.xlsx', index=False)
        df_unique_mapping = self.df.join(milestone_content_uniquenumber.set_index('milestone_number')['unique_number'],
                                         on='milestone')
        return df_unique_mapping

    def gender_info_mapping(self):  # gender_info_mapping only
        gender_set = pd.read_excel('C:/Users/홍예지/Desktop/프릭스헬스케어/raw/milestone_gender.xlsx', index=False)
        gender_info_mapping = self.df.join(gender_set.set_index('person_id')['gender'], on='person')
        return gender_info_mapping

    def unique_number_gender_info_mapping(self):  # unique_number + gender_info mapping
        gender_set = pd.read_excel('C:/Users/홍예지/Desktop/프릭스헬스케어/raw/milestone_gender.xlsx', index=False)
        df_unique_mapping = self.unique_number_mapping()
        self.df = df_unique_mapping.join(gender_set.set_index('person_id')['gender'], on='person')
        return self.df

    def sort_by_date_unique_number(self):  # 정렬
        self.df = self.df.sort_values(by=["unique_number", "done_day_after_birth"], ascending=[True, True])
        return self.df

    def event(self):  # event_numbering
        self.df['event'] = np.where(self.df['grade'] != 4, 0, 1)
        return self.df

    def done_day_after_birth_range(self, df='None', num_1=1, num_2=7878):  # done_day_after_birth 구간 지정
        condition_1 = df['done_day_after_birth'] >= num_1
        condition_2 = df['done_day_after_birth'] <= num_2
        df_date_range = df[condition_1 & condition_2]
        return df_date_range

    def gender_range(self, df='None', gender_input='both'):  # 성별정보 지정
        if gender_input == 'm':
            condition_3 = df['gender'] == 'M'
            df = df[condition_3]
            return df

        if gender_input == 'f':
            condition_4 = df['gender'] == 'F'
            df = df[condition_4]
            return df

        if gender_input == 'both':
            df = df
            return df

    def get_outlier(self, temp, column='None', weight=1.5): # outlier 제거 - IQR 이용
        quantile_25 = np.percentile(temp[column].values, 25)
        quantile_75 = np.percentile(temp[column].values, 75)

        IQR = quantile_75 - quantile_25
        IQR_weight = IQR * weight

        lowest = quantile_25 - IQR_weight
        highest = quantile_75 + IQR_weight

        outlier_idx = temp[column][(temp[column].values < lowest) | (temp[column].values > highest)].index
        return outlier_idx

    def outlier_drop(self, df='None'): # outlier 제거 - IQR 이용
        range_input = df['unique_number'].max()
        temporary=pd.DataFrame()
        for i in range(range_input):
            if i == 220:
                continue
            temp = df.groupby('unique_number').get_group(i + 1).copy()
            outlier = self.get_outlier(temp, 'done_day_after_birth', weight=1.5)
            temp.drop(outlier, axis=0, inplace=True)
            temporary=temporary.append(temp)
        self.df = temporary
        return self.df

#================================================================================================================

from lifelines import KaplanMeierFitter
from matplotlib import pyplot as plt
import matplotlib.pyplot as plt
from lifelines.plotting import qq_plot

class AverageDate: # kaplan-meier 사용 평균일자 산출 클래스
    def __init__(self, read):
        self.data_df = pd.read_excel(read)  # 데이터 프레임

        #최종 데이터 프레임에 저장
        self.milestone_number = []  #마일스톤 고유넘버
        self.mid_time = []  #마일스톤 별 done_day_after_birth 중간값
        self.start_date = []    #마일스톤 별 처음 시작 날짜
        self.due_date = []      #마일스톤 별 끝 날짜
        self.std_date = []      #마일스톤 별 표준편차

    def survival_function(self, temp):    #생존함수

        kmf = KaplanMeierFitter()
        kmf.fit_left_censoring(temp['done_day_after_birth'], temp['event'])

        mid_st = kmf.median_survival_time_
        mid_avg = self.mid_time.append(mid_st)
        # print(kmf.median_survival_time_)#중간값
        # plot = kmf.plot_survival_function() #plot 그려보고 싶을 때
        # plot.set_xlabel('done_day_after_birth')
        # plot.set_ylabel('grade_3')
        # plt.show()

        return mid_avg

    def survival_function_all(self):    #전체 done_생존함수 돌리기 시작
        range_input = self.data_df['unique_number'].max()

        for i in range(range_input):
            if i == 220:
                continue

            temp = self.data_df.groupby('unique_number').get_group(i + 1).copy()
            self.start_date.append(temp['done_day_after_birth'].min())
            self.due_date.append(temp['done_day_after_birth'].max())
            self.std_date.append(temp['done_day_after_birth'].std())

            survival = self.survival_function(temp)
            self.milestone_number.append(i+1)

        return survival

    def output_toExcel(self):   #엑셀파일 내보내기
        result_df = pd.DataFrame({"milestone_number": self.milestone_number, "day_average": self.mid_time, "std_date": self.std_date,
                                  "start_date" :self.start_date, "due_date": self.due_date})
        # print(result_df)  # 데이터 프레임 생성, 출력
        save_frame = 'C:/Users/홍예지/Desktop/프릭스헬스케어/final_test/result_test/survival_average_female_ver2.xlsx'  # 엑셀 파일 저장 위치
        result_df.to_excel(save_frame, index=False)

# ★save_frame : 결과 값 저장할 파일경로 지정해주세요.

#==============================================================================================================


class Update: # 업데이트 클래스
    def __init__(self, ori, new):
        self.ori_df = pd.read_excel(ori)    #원래 있던 파일
        self.new_df = pd.read_excel(new)    #새로운 파일

    def mergeDF(self):
        mergeDF = pd.concat([self.ori_df, self.new_df])
        mergeDF = mergeDF.sort_values(by=["milestone_unique", "done_day_after_birth"], ascending=True)

        save_frame = 'D:/test.xlsx'
        mergeDF.to_excel(save_frame, index = False)
        return mergeDF

# ★save_frame : 결과 값 저장할 파일경로 지정해주세요.

#==============================================================================================================

from openpyxl import load_workbook
import os
import scipy.stats as stats


def final_result(load_file_name): # 최종 결과 function
    load_wb = load_workbook(os.path.join(os.getcwd(), load_file_name))
    load_ws = load_wb['Sheet1']

    milestone = int(input("====> 마일스톤 넘버 입력 : "))
    day = int(input("====> 일자 입력 : "))

    print(milestone, "번의 평균 가능 일자 :", load_ws.cell(milestone + 1, 2).value, "일")
    average = load_ws.cell(milestone + 1, 2).value # average : 평균 담는 변수
    std = load_ws.cell(milestone + 1, 3).value # std : 표준편차 담는 변수
    rv = stats.norm(average, std) # rv : 평균과 표준편차를 이용하여 계산한 정규분포 담을 변수
    y3 = rv.cdf(day) # y3 : 일자에 따른 백분율 계산 값 담을 변수

    print(day, "일은 상위", "%.2f%%" % (y3 * 100), "입니다.")
    return

#==============================================================================================================

# 호출
'''
# <data_preprocessing>
new_set=Data_preprocessing('C:/Users/홍예지/Desktop/프릭스헬스케어/work/milestone.xlsx') # ★input 파일 경로 지정해주세요.
new_set.duplicates()
new_set.delete_zero_time()
new_set.gender_info_mapping()
new_set.unique_number_gender_info_mapping()
new_set.sort_by_date_unique_number()
tp=new_set.event()
tp_2=new_set.done_day_after_birth_range(tp,21,7878) #21 to 7878 까지 (전체 셋에서 유효done_day_after_birth로 취급범위)
tp_3=new_set.gender_range(tp_2,'m')
tp_4=new_set.outlier_drop(tp_3)
tp_4.to_excel('C:/Users/홍예지/Desktop/프릭스헬스케어/final_test/result_test/data_preprocessing_final_m_ver.xlsx', index=False) # ★결과 값 저장할 파일경로 지정해주세요.


'''
'''
# <data_survival_average>
test = AverageDate('C:/Users/홍예지/Desktop/프릭스헬스케어/final_test/result_test/data_preprocessing_final_f_ver.xlsx') # ★input 파일 경로 지정해주세요.
test.survival_function_all()
test.output_toExcel()
'''

# <result>
load_file_name= 'C:/Users/홍예지/Desktop/프릭스헬스케어/final_test/survival_average.xlsx' # ★input 파일 경로 지정해주세요.
final_result(load_file_name)

